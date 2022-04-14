from openpyxl import load_workbook

#https://letscode.com.br/blog/aprenda-a-integrar-python-e-excel
#https://openpyxl.readthedocs.io/en/stable/tutorial.html

#caminho = 'AULAS-EAD-REGISTRO-DE-ATIVIDADE.xlsx'
caminho = 'Aprendendoalerplanilhas.xlsx'
arquivo_excel = load_workbook(caminho, data_only=True)

planilha = arquivo_excel.active

conteudo = planilha["B5"]
planilha.cell(row=7, column=4, value='Teste euris') #row = linha; column = coluna; value = valor que está sendo acessado
arquivo_excel.save("Aprendendoalerplanilhas.xlsx")
print(conteudo.value)
