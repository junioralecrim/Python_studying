from openpyxl import load_workbook

#https://letscode.com.br/blog/aprenda-a-integrar-python-e-excel
#https://openpyxl.readthedocs.io/en/stable/tutorial.html

#caminho = 'AULAS-EAD-REGISTRO-DE-ATIVIDADE.xlsx'

#carregando o arquivo e selecionando uma página
caminho = 'Aprendendoalerplanilhas.xlsx'
arquivo_excel = load_workbook(caminho, data_only=True)

planilha = arquivo_excel.active

conteudo = planilha["B5"] #pelo oq eu entendi, aqui ele está fazendo o acesso de um jeito específico. Que seria acessando diretamente a data
#do primeiro range de alunos
print(conteudo.value)

# Tive uma idéia de como fazer!
# Pensei em criar um dicionário com cada aluno e cada um vai ter presenças e faltas, com base na idéia que o Daniel deu em sala.
# Eu vou rodar a coluna com as datas (B) e vou procurando pelo padrão de data. Após ser verdadeiro pro padrão estabelecido, eu vou
# rodando o outro for pra fazer a coleta de aluno. Vou capturar dois valores (atenção que isso pode ser feito dentro de uma função),
# o nome do aluno e se ele está presente ou não

presença = 1


#aqui eu estou acessando os dados da planilha e imprimindo na tela.
for i in range(7, 17):
    aluno = planilha.cell(row=i, column=2) #row = linha; column = coluna; value = valor que está sendo acessado
    arquivo_excel.save("Aprendendoalerplanilhas.xlsx") #ainda não entendi pra que serve isso??
    print(aluno.value)

    #o tratamento de dados vai ser feito aqui.
    if(i == 16):
        for j in range(7, 17):
            aluno = planilha.cell(row=j, column=3)  # row = linha; column = coluna; value = valor que está sendo acessado
            arquivo_excel.save("Aprendendoalerplanilhas.xlsx")  # ainda não entendi pra que serve isso??
            print(aluno.value)


