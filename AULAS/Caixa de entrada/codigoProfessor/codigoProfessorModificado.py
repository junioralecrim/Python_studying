from openpyxl import load_workbook

#https://letscode.com.br/blog/aprenda-a-integrar-python-e-excel
#https://openpyxl.readthedocs.io/en/stable/tutorial.html

#caminho = 'AULAS-EAD-REGISTRO-DE-ATIVIDADE.xlsx'
caminho = 'Aprendendoalerplanilhas.xlsx'
arquivo_excel = load_workbook(caminho, data_only=True)

planilha = arquivo_excel.active


#planilha.cell(row=7, column=4, value='Teste euris')
#arquivo_excel.save("Aprendendoalerplanilhas.xlsx")
alunos = []
for i in range(7, 17):
    conteudo = planilha[f'B{i}']
    alunos.append(conteudo.value)

resultadoAlunosSim = {}
resultadoAlunosNao = {}

for a in alunos:
    somasim = 0
    somanao = 0
    for i in range(7, 36):
        celulaAluno = planilha[f'B{i}']
        celulaResposta = planilha[f'C{i}']
        if celulaAluno.value == a and celulaResposta.value == 'SIM':
            somasim += 1
            resultadoAlunosSim[a] = somasim
        elif (celulaAluno.value == a) and (celulaResposta.value == 'NÃO' or celulaResposta.value == 'NAO'):
            somanao += 1
            resultadoAlunosNao[a] = somanao
print("SIM", resultadoAlunosSim)
print("NÃO", resultadoAlunosNao)


print("\n")
teste = 0

for a in alunos:
    somanao = 0
    for i in range(21, 28):
        celulaAluno = planilha[f'B{i}']
        if a == celulaAluno.value:
            break
        #rodar todos e não achar correspondência
        if i == 27:
            print(f"Teste: {celulaAluno.value} {a}")
            teste += 1
            somanao += 1
            print(f"\n\nTeste de controle: \nResultado Aluno Não: {resultadoAlunosNao}\nResultado variavel somanao:{somanao}")
            resultadoAlunosNao[a] = somanao

