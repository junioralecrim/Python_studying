#https://docs.google.com/spreadsheets/d/1gw1S-d-Ywz8RuXy2D4yXWuerC5-Cb2g7Cs7klUltz_M/edit?usp=sharing


#dados que eu quero: Data, professor, horário e conteúdo ministrado


from openpyxl import load_workbook
aux = 0
conteudoAula = []
horaAula = []
#https://letscode.com.br/blog/aprenda-a-integrar-python-e-excel
#https://openpyxl.readthedocs.io/en/stable/tutorial.html

#caminho = 'AULAS-EAD-REGISTRO-DE-ATIVIDADE.xlsx'


caminho = 'Aula de Algoritmos 2 - AULAS EAD REGISTRO DE ATIVIDADE.xlsx'
arquivo_excel = load_workbook(caminho, data_only=True)

planilha = arquivo_excel.active
conteudo = planilha['B1']
celula = []
celula.append(conteudo.value)

#planilha.cell(row=7, column=4) #row = linha, column = coluna
#arquivo_excel.save("Aprendendoalerplanilhas.xlsx")

alunos = []
for i in range(6, 232):
   conteudo = planilha[f'B{i}']
   professorSN = str(conteudo.value)
   professorSN = professorSN.lower()

   if(professorSN == 'professor'):
       conteudo = planilha[f'C{i}']
       nome = str(conteudo.value)
       nome = nome.lower()
       if(nome == 'euristenho júnior') or (nome == 'euristenho jr.'):
           conteudo = planilha[f'C{i+1}']
           horaAula.append(conteudo.value)
           conteudo = planilha[f'C{i + 4}']
           conteudoAula.append(conteudo.value)


print(f"HORA AULA\n{horaAula}")
print(f"\nCONTEUDO AULA\n{conteudoAula}")

#https://giovannireisnunes.wordpress.com/2021/02/19/manipulando-xlsx-em-python/